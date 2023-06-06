import os
import glob
import pandas as pd
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from dotenv import load_dotenv

load_dotenv()

folder = os.environ.get('FOLDER')
dynamic_folder = os.environ.get('DYNAMIC_FOLDER')

# Obtener archivos
excel_files = glob.glob(os.path.join(folder, '*.xlsx'))

# Ordenar archivos por fecha de modificación
excel_files.sort(key=os.path.getmtime, reverse=True)

# Verificar si hay archivos en la lista
if excel_files:

    # Seleccionar el archivo más nuevo de la lista
    ruta_archivo = excel_files[0]
    try:
        df = pd.read_excel(ruta_archivo)
        pivot_table = df.groupby('Solicitante').size().reset_index(name='Count')  

        # Ordenar la tabla de mayor a menor según la columna 'num'
        pivot_table = pivot_table.sort_values(by='Count', ascending=False)

        # Continúa con el código para trabajar con la tabla pivote ordenada y sin la celda vacía
        from openpyxl import Workbook
        from openpyxl.utils.dataframe import dataframe_to_rows

        # Obtener la fecha de inicio de la última fila en la columna 'Fecha inicio'
        ultima_fecha_inicio = df['Fecha inicio'].iloc[-1]

        # Crear un nombre para el nuevo archivo Excel
        nombre_archivo = f"Reporte_{ultima_fecha_inicio.strftime('%Y%m%d%H%M%S')}.xlsx"

        # Crear un nuevo libro de trabajo y seleccionar la hoja activa
        workbook = Workbook()
        sheet = workbook.active

        # Escribir la tabla pivote en la hoja de cálculo
        for i, (index, row) in enumerate(pivot_table.iterrows(), start=1):
            sheet[f'A{i}'] = row['Solicitante']
            sheet[f'B{i}'] = row['Count'] 

        # Alinear los datos en la hoja de cálculo
        for column in sheet.columns:
            max_length = 0
            column = [cell for cell in column]
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass
            adjusted_width = (max_length + 2)
            sheet.column_dimensions[get_column_letter(column[0].column)].width = adjusted_width
            if column[0].column == 2:  # Ajustar ancho de la columna B
                sheet.column_dimensions[get_column_letter(column[0].column)].width = adjusted_width * 2

        # Guardar el archivo Excel
        workbook.save(os.path.join(dynamic_folder, nombre_archivo))

        print(f"Se ha creado el archivo '{nombre_archivo}' exitosamente en '{dynamic_folder}' usando '{ruta_archivo}'")
    except Exception as e:
        print("Se produjo un error:", e)
else:
    print("No se encontraron archivos Excel en la carpeta.")