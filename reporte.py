import os
import glob
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from dotenv import load_dotenv
import tkinter as tk
from tkinter import filedialog, messagebox
import time

# pyinstaller --onefile name.py / pip install, pyinstaller.

load_dotenv()

def get_newest_file(folder):
    excel_files = glob.glob(os.path.join(folder, '*.xlsx'))
    excel_files.sort(key=os.path.getmtime, reverse=True)
    if excel_files:
        return excel_files[0]
    return None

def generate_pivot_table(df, column):
    pivot_table = df.groupby(column).size().reset_index(name='Count')
    pivot_table = pivot_table.sort_values(by='Count', ascending=False)
    return pivot_table

def generate_excel_report(pivot_table_titulo, pivot_table_solicitante, pivot_table_area_solicitante, pivot_table_clasificacion, output_folder, df):
    workbook = Workbook()
    sheet = workbook.active
    
    styles(sheet['A1'], 'Título', "249B22", "FFFFFF", True)
    styles(sheet['B1'], '#', "249B22", "FFFFFF", True)
    styles(sheet['D1'], 'Solicitante', "249B22", "FFFFFF", True)
    styles(sheet['E1'], '#', "249B22", "FFFFFF", True)
    styles(sheet['G1'], 'Área del Solicitante', "249B22", "FFFFFF", True)
    styles(sheet['H1'], '#', "249B22", "FFFFFF", True)
    styles(sheet['J1'], 'Clasificación', "249B22", "FFFFFF", True)
    styles(sheet['K1'], '#', "249B22", "FFFFFF", True)

    sheet.column_dimensions['A'].width = 45
    sheet.column_dimensions['D'].width = 45
    sheet.column_dimensions['G'].width = 45
    sheet.column_dimensions['J'].width = 45

    sheet.move_range(f"A2:B{pivot_table_titulo.shape[0] + 1}", rows=1)
    sheet.move_range(f"D2:E{pivot_table_solicitante.shape[0] + 1}", rows=1)
    sheet.move_range(f"G2:H{pivot_table_area_solicitante.shape[0] + 1}", rows=1)
    sheet.move_range(f"J2:K{pivot_table_clasificacion.shape[0] + 1}", rows=1)
    
    for i, (index, row) in enumerate(pivot_table_titulo.iterrows(), start=2):
        sheet[f'A{i}'] = row['Título']
        sheet[f'B{i}'] = row['Count']

    for i, (index, row) in enumerate(pivot_table_solicitante.iterrows(), start=2):
        sheet[f'D{i}'] = row['Solicitante']
        sheet[f'E{i}'] = row['Count']

    for i, (index, row) in enumerate(pivot_table_area_solicitante.iterrows(), start=2):
        sheet[f'G{i}'] = row['Área del Solicitante']
        sheet[f'H{i}'] = row['Count']
    for i, (index, row) in enumerate(pivot_table_clasificacion.iterrows(), start=2):
        sheet[f'J{i}'] = row['Clasificacion']
        sheet[f'K{i}'] = row['Count']

    for column in sheet.columns:
        max_length = 0
        column = [cell for cell in column]
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        
    last_date_row_Fecha_Inicio = df['Fecha inicio'].iloc[-1]
    newfile_name = f"Reporte_{last_date_row_Fecha_Inicio.strftime('%Y%m%d')}.xlsx"
    file_path = os.path.join(output_folder, newfile_name)
    workbook.save(file_path)
    return file_path


def styles(cell, value, fill_color, font_color, is_bold=False):
    cell.value = value
    cell.fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")
    cell.font = Font(color=font_color, bold=is_bold)

def select_excel_file():
    file_path = filedialog.askopenfilename(filetypes=[("Excel Files", "*.xlsx")])
    if file_path:
        try:
            df = pd.read_excel(file_path)
            pivot_table_titulo = generate_pivot_table(df, 'Título')
            pivot_table_solicitante = generate_pivot_table(df, 'Solicitante')
            pivot_table_area_solicitante = generate_pivot_table(df, 'Área del Solicitante')
            pivot_table_clasificacion = generate_pivot_table(df, 'Clasificacion')
            output_folder = filedialog.askdirectory()
            if output_folder:
                generated_file = generate_excel_report(pivot_table_titulo, pivot_table_solicitante, pivot_table_area_solicitante, pivot_table_clasificacion, output_folder, df)
                messagebox.showinfo("Éxito", f"Se ha creado el archivo con éxito:\n{generated_file}")
            else:
                messagebox.showerror("Error", "No se ha seleccionado una carpeta de salida.")
        except Exception as e:
            messagebox.showerror("Error", f"Se produjo un error:\n{str(e)}")
    else:
        messagebox.showerror("Error", "No se ha seleccionado un archivo Excel.")

def main():
    root = tk.Tk()
    root.title("Generador de Excel")
    root.geometry("300x200")
    
    select_button = tk.Button(root, text="Seleccionar archivo", command=select_excel_file)
    select_button.pack(pady=50)
    
    root.mainloop()

if __name__ == "__main__":
    main()
