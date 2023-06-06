import os
import glob
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from dotenv import load_dotenv
import time

load_dotenv()

def get_newest_file(folder):
    excel_files = glob.glob(os.path.join(folder, '*.xlsx'))
    excel_files.sort(key=os.path.getmtime, reverse=True)
    if excel_files:
        return excel_files[0]
    return None

def generate_pivot_table(df, column):
    pivot_table = df.groupby(column).size().reset_index(name='Count')
    pivot_table = pivot_table.sort_values(by='Count', ascending=False)
    return pivot_table

def generate_excel_report(pivot_table_titulo, pivot_table_solicitante, pivot_table_area_solicitante, pivot_table_clasificacion, output_folder, df):
    workbook = Workbook()
    sheet = workbook.active
    
    styles(sheet['A1'], 'Título', "249B22", "FFFFFF", True)
    styles(sheet['B1'], '#', "249B22", "FFFFFF", True)
    styles(sheet['D1'], 'Solicitante', "249B22", "FFFFFF", True)
    styles(sheet['E1'], '#', "249B22", "FFFFFF", True)
    styles(sheet['G1'], 'Área del Solicitante', "249B22", "FFFFFF", True)
    styles(sheet['H1'], '#', "249B22", "FFFFFF", True)
    styles(sheet['J1'], 'Clasificación', "249B22", "FFFFFF", True)
    styles(sheet['K1'], '#', "249B22", "FFFFFF", True)

    sheet.column_dimensions['A'].width = 45
    sheet.column_dimensions['D'].width = 45
    sheet.column_dimensions['G'].width = 45
    sheet.column_dimensions['J'].width = 45

    sheet.move_range(f"A2:B{pivot_table_titulo.shape[0] + 1}", rows=1)
    sheet.move_range(f"D2:E{pivot_table_solicitante.shape[0] + 1}", rows=1)
    sheet.move_range(f"G2:H{pivot_table_area_solicitante.shape[0] + 1}", rows=1)
    sheet.move_range(f"J2:K{pivot_table_clasificacion.shape[0] + 1}", rows=1)
    
    for i, (index, row) in enumerate(pivot_table_titulo.iterrows(), start=2):
        sheet[f'A{i}'] = row['Título']
        sheet[f'B{i}'] = row['Count']

    for i, (index, row) in enumerate(pivot_table_solicitante.iterrows(), start=2):
        sheet[f'D{i}'] = row['Solicitante']
        sheet[f'E{i}'] = row['Count']

    for i, (index, row) in enumerate(pivot_table_area_solicitante.iterrows(), start=2):
        sheet[f'G{i}'] = row['Área del Solicitante']
        sheet[f'H{i}'] = row['Count']
    for i, (index, row) in enumerate(pivot_table_clasificacion.iterrows(), start=2):
        sheet[f'J{i}'] = row['Clasificacion']
        sheet[f'K{i}'] = row['Count']

    for column in sheet.columns:
        max_length = 0
        column = [cell for cell in column]
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        
    last_date_row_Fecha_Inicio = df['Fecha inicio'].iloc[-1]
    newfile_name = f"Reporte_{last_date_row_Fecha_Inicio.strftime('%Y%m%d')}.xlsx"
    file_path = os.path.join(output_folder, newfile_name)
    workbook.save(file_path)
    return file_path


def styles(cell, value, fill_color, font_color, is_bold=False):
    cell.value = value
    cell.fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")
    cell.font = Font(color=font_color, bold=is_bold)

def main():
    start_time = time.time()
    folder = os.environ.get('FOLDER')
    dynamic_folder = os.environ.get('DYNAMIC_FOLDER')

    newest_file = get_newest_file(folder)

    if newest_file:
        try:
            df = pd.read_excel(newest_file)
            pivot_table_titulo = generate_pivot_table(df, 'Título')
            pivot_table_solicitante = generate_pivot_table(df, 'Solicitante')
            pivot_table_area_solicitante = generate_pivot_table(df, 'Área del Solicitante')
            pivot_table_clasificacion = generate_pivot_table(df, 'Clasificacion')
            generated_file = generate_excel_report(pivot_table_titulo, pivot_table_solicitante, pivot_table_area_solicitante, pivot_table_clasificacion, dynamic_folder, df)
            end_time = time.time()
            execution_time = end_time - start_time
            print(f"Se ha creado el archivo con éxito con un tiempo de ejecución de: ", execution_time, " segundos.")

        except Exception as e:
            print("Se produjo un error:", e)
            print("Tiempo de ejecución:", execution_time, "segundos")
    else:
        print("No se encontraron archivos Excel en la carpeta.")

if __name__ == "__main__":
    main()